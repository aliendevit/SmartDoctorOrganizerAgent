# new_client_tab.py
import os
import json
from PyQt5 import QtWidgets, QtCore, QtGui
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    print("openpyxl is required. Please install it using 'pip install openpyxl'.")

# Define the Excel file name that will store client names.
EXCEL_FILE = "clients.xlsx"

class NewClientTab(QtWidgets.QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setup_ui()

    def setup_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(15, 15, 15, 15)
        layout.setSpacing(10)

        # Header label
        header_label = QtWidgets.QLabel("Add New Client Manually")
        header_label.setAlignment(QtCore.Qt.AlignCenter)
        header_label.setFont(QtGui.QFont("Arial", 16, QtGui.QFont.Bold))
        layout.addWidget(header_label)

        # Create form fields for client data
        form_layout = QtWidgets.QFormLayout()
        self.name_edit = QtWidgets.QLineEdit()
        form_layout.addRow("Name:", self.name_edit)
        self.age_edit = QtWidgets.QLineEdit()
        form_layout.addRow("Age:", self.age_edit)
        self.symptoms_edit = QtWidgets.QLineEdit()
        form_layout.addRow("Symptoms (comma-separated):", self.symptoms_edit)
        self.notes_edit = QtWidgets.QTextEdit()
        self.notes_edit.setFixedHeight(80)
        form_layout.addRow("Notes:", self.notes_edit)
        layout.addLayout(form_layout)

        # Button 1: Save as PDF & JSON
        self.save_pdf_json_button = QtWidgets.QPushButton("Save as PDF & JSON")
        self.save_pdf_json_button.setFixedHeight(40)
        self.save_pdf_json_button.clicked.connect(self.save_client_to_pdf_json)
        layout.addWidget(self.save_pdf_json_button)

        # Button 2: Add Client Name to Excel
        self.add_to_excel_button = QtWidgets.QPushButton("Add Client Name to Excel")
        self.add_to_excel_button.setFixedHeight(40)
        self.add_to_excel_button.clicked.connect(self.append_client_name_to_excel)
        layout.addWidget(self.add_to_excel_button)

        layout.addStretch()

    def get_client_data(self):
        """
        Collect data from form fields and return it as a dictionary.
        If the name is missing, show a warning and return None.
        """
        name = self.name_edit.text().strip()
        age = self.age_edit.text().strip()
        symptoms = self.symptoms_edit.text().strip()
        notes = self.notes_edit.toPlainText().strip()

        if not name:
            QtWidgets.QMessageBox.warning(self, "Input Error", "Please enter the client's name.")
            return None

        client_data = {
            "Name": name,
            "Age": age if age else "N/A",
            "Symptoms": [s.strip() for s in symptoms.split(",")] if symptoms else [],
            "Notes": notes if notes else ""
        }
        return client_data

    def save_client_to_pdf_json(self):
        """
        Save the full client data as both a PDF and a JSON file.
        Both files will use the client's name as the base file name.
        """
        client_data = self.get_client_data()
        if client_data is None:
            return

        # Sanitize the client name for file naming.
        client_name = client_data["Name"]
        client_name_sanitized = "".join(c for c in client_name if c.isalnum() or c in (' ', '_')).strip().replace(" ", "_")
        if not client_name_sanitized:
            client_name_sanitized = "client"

        # Ask the user for a directory where to save the files
        directory = QtWidgets.QFileDialog.getExistingDirectory(self, "Select Directory to Save Files", "")
        if not directory:
            return

        # Create file paths
        pdf_file_path = os.path.join(directory, f"{client_name_sanitized}.pdf")
        json_file_path = os.path.join(directory, f"{client_name_sanitized}.json")

        # Save JSON file
        try:
            with open(json_file_path, "w") as f:
                json.dump(client_data, f, indent=4)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Save Error", f"Error saving JSON: {e}")
            return

        # Save PDF file using ReportLab
        try:
            doc = SimpleDocTemplate(pdf_file_path, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()
            title = Paragraph("New Client Data", styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 12))
            table_data = [["Field", "Value"]]
            for field, value in client_data.items():
                if isinstance(value, list):
                    value = ", ".join(value)
                table_data.append([field, str(value)])
            t = Table(table_data, colWidths=[150, 350])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(t)
            doc.build(elements)
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Save Error", f"Error saving PDF: {e}")
            return

        QtWidgets.QMessageBox.information(self, "Success",
            f"Client data saved successfully!\n\nPDF: {pdf_file_path}\nJSON: {json_file_path}")

    def append_client_name_to_excel(self):
        """
        Append only the client's name to an Excel file that accumulates all client names.
        Each time this button is pressed, the client's name is added as a new row.
        """
        client_data = self.get_client_data()
        if client_data is None:
            return

        # We'll use only the client's name.
        client_name = client_data["Name"]

        try:
            from openpyxl import load_workbook, Workbook
        except ImportError:
            QtWidgets.QMessageBox.critical(self, "Dependency Error", "openpyxl is not installed. Use 'pip install openpyxl'.")
            return

        # Check if the Excel file exists; if not, create it with a header row.
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Name"])  # header row

        # Append the new client name
        ws.append([client_name])
        try:
            wb.save(EXCEL_FILE)
            QtWidgets.QMessageBox.information(self, "Success", f"Client name added to Excel file:\n{EXCEL_FILE}")
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, "Save Error", f"Error saving to Excel: {e}")

# For independent testing, uncomment below:
# if __name__ == "__main__":
#     import sys
#     app = QtWidgets.QApplication(sys.argv)
#     widget = NewClientTab()
#     widget.show()
#     sys.exit(app.exec_())
