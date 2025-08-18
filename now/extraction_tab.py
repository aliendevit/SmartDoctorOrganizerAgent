import os
import re
import json
import spacy
from datetime import datetime, timedelta
from PyQt5 import QtWidgets, QtCore, QtGui
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

try:
    from dateutil.relativedelta import relativedelta
    DATEUTIL_AVAILABLE = True
except ImportError:
    DATEUTIL_AVAILABLE = False

nlp = spacy.load("en_core_web_sm")

import speech_recognition as sr

# -------------------- Voice Input Widget --------------------
class VoiceInputWidget(QtWidgets.QWidget):
    textReady = QtCore.pyqtSignal(str)

    def __init__(self, parent=None, language="en-US"):
        super().__init__(parent)
        self.language = language
        self.setup_ui()

    def setup_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        self.voice_button = QtWidgets.QPushButton()
        self.voice_button.setFont(QtGui.QFont("Segoe UI", 14))
        self.voice_button.clicked.connect(self.start_voice_input)
        layout.addWidget(self.voice_button)
        self.setLayout(layout)
        self.retranslateUi()

    def retranslateUi(self):
        self.voice_button.setText(self.tr("Voice Input ({0})").format(self.language))

    def start_voice_input(self):
        self.voice_button.setText(self.tr("Listening... ({0})").format(self.language))
        QtWidgets.QApplication.processEvents()
        r = sr.Recognizer()
        try:
            with sr.Microphone() as source:
                r.adjust_for_ambient_noise(source, duration=0.5)
                audio = r.listen(source, timeout=7)
                text = r.recognize_google(audio, language=self.language)
                self.textReady.emit(text)
                self.voice_button.setText(self.tr("Voice Input ({0})").format(self.language))
        except sr.WaitTimeoutError:
            QtWidgets.QMessageBox.warning(self, self.tr("Voice Input Error"),
                                          self.tr("Listening timed out. Please try again."))
            self.voice_button.setText(self.tr("Voice Input ({0})").format(self.language))
        except sr.UnknownValueError:
            QtWidgets.QMessageBox.warning(self, self.tr("Voice Input Error"),
                                          self.tr("Could not understand the audio. Please speak clearly."))
            self.voice_button.setText(self.tr("Voice Input ({0})").format(self.language))
        except sr.RequestError as e:
            QtWidgets.QMessageBox.warning(self, self.tr("Voice Input Error"),
                                          self.tr("Could not request results; {0}").format(e))
            self.voice_button.setText(self.tr("Voice Input ({0})").format(self.language))
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, self.tr("Voice Input Error"),
                                          self.tr("An unexpected error occurred: {0}").format(e))
            self.voice_button.setText(self.tr("Voice Input ({0})").format(self.language))

# Dummy implementation of parse_patient_info.
def parse_patient_info(text):
    # Replace with your actual logic.
    return {
        "Name": "John Doe",
        "Age": 30,
        "Symptoms": ["cough", "fever"],
        "Notes": "Patient shows mild symptoms.",
        "Date": datetime.today().strftime("%d-%m-%Y"),
        "Appointment Date": "Not Specified",
        "Appointment Time": "Not Specified",
        "Summary": "Patient exhibits mild symptoms.",
        "Follow-Up Date": (datetime.today() + timedelta(days=7)).strftime("%d-%m-%Y")
    }

# Dummy implementation of generate_pdf_report.
def generate_pdf_report(data, file_path):
    doc = SimpleDocTemplate(file_path, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []
    client_name = data.get("Name", "Unknown")
    title = Paragraph(f"Patient Report: {client_name}", styles["Title"])
    elements.append(title)
    elements.append(Spacer(1, 12))
    summary = data.get("Summary", "No summary available.")
    summary_paragraph = Paragraph(f"<b>Summary:</b><br/>{summary}", styles["BodyText"])
    elements.append(summary_paragraph)
    elements.append(Spacer(1, 12))
    header_data = [Paragraph("<b>Field</b>", styles["BodyText"]), Paragraph("<b>Value</b>", styles["BodyText"])]
    details = [header_data]
    fields = [
        ("Age", data.get("Age", "N/A")),
        ("Symptoms", ", ".join(data.get("Symptoms", []))),
        ("Notes", data.get("Notes", "N/A")),
        ("General Date", data.get("Date", "Not Specified")),
        ("Appointment Date", data.get("Appointment Date", "Not Specified")),
        ("Appointment Time", data.get("Appointment Time", "Not Specified")),
        ("Follow-Up Date", data.get("Follow-Up Date", "Not Specified"))
    ]
    for field, value in fields:
        details.append([Paragraph(field, styles["BodyText"]), Paragraph(str(value), styles["BodyText"])])
    table = Table(details, colWidths=[150, 350])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#008080')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    elements.append(table)
    doc.build(elements)

def generate_summary(text):
    sentences = re.split(r'(?<=[.!?])\s+', text.strip())
    if len(sentences) >= 2:
        return " ".join(sentences[:2])
    elif sentences:
        return sentences[0]
    else:
        return ""

# -------------------- ExtractionTab Widget --------------------
class ExtractionTab(QtWidgets.QWidget):
    dataProcessed = QtCore.pyqtSignal(dict)
    appointmentProcessed = QtCore.pyqtSignal(dict)
    switchToAppointments = QtCore.pyqtSignal(str)

    def tr(self, text):
        from translation_helper import tr
        return tr(text)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.automation_paused = False
        self.setup_ui()

    def setup_ui(self):
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        # Pause Voice button
        self.pause_voice_btn = QtWidgets.QPushButton(self.tr("Pause Voice"))
        self.pause_voice_btn.setFont(QtGui.QFont("Segoe UI", 14))
        self.pause_voice_btn.clicked.connect(self.toggle_voice_input)
        layout.addWidget(self.pause_voice_btn)

        # Header label
        self.header_label = QtWidgets.QLabel(self.tr("Data Extraction"))
        self.header_label.setAlignment(QtCore.Qt.AlignCenter)
        self.header_label.setFont(QtGui.QFont("Segoe UI", 22, QtGui.QFont.Bold))
        layout.addWidget(self.header_label)

        # Prompt label
        self.prompt_label = QtWidgets.QLabel(self.tr("Enter patient information and AI will analyze it:"))
        self.prompt_label.setFont(QtGui.QFont("Segoe UI", 16))
        layout.addWidget(self.prompt_label)

        # Horizontal layout: text input and voice input widget
        input_layout = QtWidgets.QHBoxLayout()
        self.input_text = QtWidgets.QTextEdit()
        self.input_text.setMinimumHeight(120)
        input_layout.addWidget(self.input_text)
        self.voice_widget_ar = VoiceInputWidget(language="ar-SA")
        self.voice_widget_ar.textReady.connect(lambda text: self.input_text.setPlainText(text))
        input_layout.addWidget(self.voice_widget_ar)
        layout.addLayout(input_layout)

        # Buttons layout for test data and processing
        self.load_test_button = QtWidgets.QPushButton(self.tr("Load Test Data"))
        self.load_test_button.clicked.connect(self.load_test_data)
        self.process_button = QtWidgets.QPushButton(self.tr("Process Input"))
        self.process_button.clicked.connect(self.delayed_process_input)
        buttons_layout = QtWidgets.QHBoxLayout()
        buttons_layout.addWidget(self.load_test_button)
        buttons_layout.addWidget(self.process_button)
        layout.addLayout(buttons_layout)

        # Table for displaying parsed data
        self.tableWidget = QtWidgets.QTableWidget()
        self.tableWidget.setColumnCount(2)
        self.tableWidget.setHorizontalHeaderLabels([self.tr("Field"), self.tr("Value")])
        self.tableWidget.horizontalHeader().setStretchLastSection(True)
        self.tableWidget.setEditTriggers(QtWidgets.QTableWidget.NoEditTriggers)
        layout.addWidget(self.tableWidget)

        # Save and automation buttons
        self.save_pdf_json_button = QtWidgets.QPushButton(self.tr("Create Report"))
        self.save_pdf_json_button.clicked.connect(self.save_data)
        self.append_excel_button = QtWidgets.QPushButton(self.tr("Append Name to Clients File"))
        self.append_excel_button.clicked.connect(self.append_client_name_to_excel)
        save_buttons_layout = QtWidgets.QHBoxLayout()
        save_buttons_layout.addWidget(self.save_pdf_json_button)
        save_buttons_layout.addWidget(self.append_excel_button)
        layout.addLayout(save_buttons_layout)

        # Automation button
        self.start_aaa_btn = QtWidgets.QPushButton(self.tr("Start AAA"))
        self.start_aaa_btn.clicked.connect(self.automate_aaa)
        automation_layout = QtWidgets.QHBoxLayout()
        automation_layout.addWidget(self.start_aaa_btn)
        layout.addLayout(automation_layout)

        # Status label
        self.status_label = QtWidgets.QLabel(self.tr("Status: Ready"))
        self.status_label.setAlignment(QtCore.Qt.AlignCenter)
        self.status_label.setFont(QtGui.QFont("Segoe UI", 14, QtGui.QFont.Bold))
        layout.addWidget(self.status_label)

        layout.addStretch()
        self.setLayout(layout)
        self.apply_styles()
        self.retranslateUi()

        # Initialize automation timers with longer delays to allow processing to complete.
        self.aaa_timer1 = QtCore.QTimer(self)
        self.aaa_timer1.setSingleShot(True)
        self.aaa_timer2 = QtCore.QTimer(self)
        self.aaa_timer2.setSingleShot(True)

    def retranslateUi(self):
        self.pause_voice_btn.setText(self.tr("Pause Voice"))
        self.header_label.setText(self.tr("Data Extraction"))
        self.prompt_label.setText(self.tr("Enter patient information and AI will analyze it:"))
        self.load_test_button.setText(self.tr("Load Test Data"))
        self.process_button.setText(self.tr("Process Input"))
        self.tableWidget.setHorizontalHeaderLabels([self.tr("Field"), self.tr("Value")])
        self.save_pdf_json_button.setText(self.tr("Create Report"))
        self.append_excel_button.setText(self.tr("Append Name to Clients File"))
        self.start_aaa_btn.setText(self.tr("Start AAA"))
        self.status_label.setText(self.tr("Status: Ready"))
        self.voice_widget_ar.retranslateUi()

    def apply_styles(self):
        style = """
            QWidget {
                background-color: #f0f0f0;
                color: #003366;
                font-family: Arial, sans-serif;
                font-size: 14px;
            }
            QLabel {
                color: #003366;
            }
            QTextEdit {
                background-color: #ffffff;
                border: 1px solid #99C2E0;
                border-radius: 8px;
                padding: 12px;
            }
            QPushButton {
                background-color: #008080;
                border: none;
                border-radius: 8px;
                padding: 10px 18px;
                color: white;
                font-size: 16px;
            }
            QPushButton:hover {
                background-color: #006666;
            }
            QPushButton#start_aaa {
                background-color: #28A745;
                font-weight: bold;
            }
            QPushButton#start_aaa:hover {
                background-color: #218838;
            }
            QTableWidget {
                background-color: #ffffff;
                border: 1px solid #E1E4E8;
                border-radius: 8px;
            }
            QHeaderView::section {
                background-color: #008080;
                padding: 10px;
                border: none;
                color: white;
                font-size: 16px;
            }
            QScrollBar:vertical {
                background: #99C2E0;
                width: 12px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical {
                background: #008080;
                min-height: 20px;
                border-radius: 6px;
            }
            QScrollBar::handle:vertical:hover {
                background: #006666;
            }
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                height: 0px;
            }
        """
        self.start_aaa_btn.setObjectName("start_aaa")
        self.setStyleSheet(style)

    # Dummy implementation for load_test_data
    def load_test_data(self):
        test_data = (
            "Patient Jane Smith, age 23, complains of cough and headache. "
            "Her appointment is scheduled for today. "
            "Her follow-up is scheduled for 22-11-2023. "
            "She has a history of asthma. "
            "Summary: Patient exhibits mild respiratory distress."
        )
        self.input_text.setPlainText(test_data)

    def delayed_process_input(self):
        self.status_label.setText(self.tr("Status: Processing input..."))
        self.thinking_msg = QtWidgets.QMessageBox(self)
        self.thinking_msg.setWindowTitle(self.tr("Processing"))
        self.thinking_msg.setText(self.tr("Generating output... Please wait."))
        self.thinking_msg.setStandardButtons(QtWidgets.QMessageBox.NoButton)
        self.thinking_msg.show()
        self.input_text.setDisabled(True)
        # Increase delay to ensure processing completes.
        QtCore.QTimer.singleShot(4000, self.process_input)

    def process_input(self):
        try:
            text = self.input_text.toPlainText()
            if not text.strip():
                QtWidgets.QMessageBox.warning(self, self.tr("Input Error"), self.tr("Please enter some text."))
                self.input_text.setDisabled(False)
                self.thinking_msg.hide()
                self.status_label.setText(self.tr("Status: Ready"))
                return
            # Dummy parsing – replace with your actual parsing logic.
            self.current_data = parse_patient_info(text)
            self.populate_table(self.current_data)
            self.dataProcessed.emit(self.current_data)
            if self.current_data.get("Appointment Date", "Not Specified") != "Not Specified":
                self.appointmentProcessed.emit(self.current_data)
            from data.data import insert_client
            insert_client(self.current_data)
            self.status_label.setText(self.tr("Status: Input processed successfully."))
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Processing Error"), self.tr("An error occurred:\n") + str(e))
            self.status_label.setText(self.tr("Status: Error processing input."))
        finally:
            self.input_text.setDisabled(False)
            self.thinking_msg.hide()

    def populate_table(self, data):
        self.tableWidget.setRowCount(0)
        for row, (field, value) in enumerate(data.items()):
            self.tableWidget.insertRow(row)
            field_item = QtWidgets.QTableWidgetItem(field)
            if isinstance(value, list):
                value = ", ".join(value)
            value_item = QtWidgets.QTableWidgetItem(str(value))
            field_item.setFont(QtGui.QFont("Segoe UI", 13))
            value_item.setFont(QtGui.QFont("Segoe UI", 13))
            self.tableWidget.setItem(row, 0, field_item)
            self.tableWidget.setItem(row, 1, value_item)

    def save_data(self):
        if not hasattr(self, "current_data") or not self.current_data:
            QtWidgets.QMessageBox.warning(self, self.tr("Save Data Error"),
                                          self.tr("Please process input before creating a report."))
            return
        self.status_label.setText(self.tr("Status: Saving report..."))
        try:
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            save_directory = os.path.join(desktop, "reports")
            if not os.path.exists(save_directory):
                os.makedirs(save_directory)
            client_name = self.current_data.get("Name", "Unknown")
            base_filename = "".join(c for c in client_name if c.isalnum() or c in (' ', '_')).replace(" ", "_")
            if not base_filename:
                base_filename = "Unknown"
            pdf_file_path = os.path.join(save_directory, f"{base_filename}_report.pdf")
            json_file_path = os.path.join(save_directory, f"{base_filename}_report.json")
            try:
                generate_pdf_report(self.current_data, pdf_file_path)
            except Exception as report_error:
                QtWidgets.QMessageBox.critical(self, self.tr("Report Creation Error"),
                                               self.tr("An error occurred while creating the report:\n") + str(report_error))
                self.status_label.setText(self.tr("Status: Error in report creation."))
                return
            with open(json_file_path, "w") as f:
                json.dump(self.current_data, f, indent=4)
            report_box = QtWidgets.QMessageBox(self)
            report_box.setWindowTitle(self.tr("Report Created"))
            report_box.setText(self.tr("Report saved as PDF and JSON:\n") + pdf_file_path + "\n" + json_file_path)
            report_box.setStandardButtons(QtWidgets.QMessageBox.Ok)
            QtCore.QTimer.singleShot(2000, report_box.accept)
            report_box.exec_()
            self.status_label.setText(self.tr("Status: Report created."))
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Save Data Error"),
                                           self.tr("An error occurred while saving reports:\n") + str(e))
            self.status_label.setText(self.tr("Status: Error saving report."))

    def append_client_name_to_excel(self):
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            QtWidgets.QMessageBox.warning(self, self.tr("Excel Error"),
                                          self.tr("openpyxl is required. Please install it via 'pip install openpyxl'."))
            return
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        excel_path = os.path.join(desktop, "clients.xlsx")
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
            ws = wb.active
        else:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "Client Name"
        client_name = self.current_data.get("Name", "Unknown")
        ws.append([client_name])
        wb.save(excel_path)
        excel_box = QtWidgets.QMessageBox(self)
        excel_box.setWindowTitle(self.tr("Excel"))
        excel_box.setText(self.tr("Client name appended to Excel!\nFile path: ") + excel_path)
        excel_box.setStandardButtons(QtWidgets.QMessageBox.Ok)
        QtCore.QTimer.singleShot(2000, excel_box.accept)
        excel_box.exec_()
        self.status_label.setText(self.tr("Status: Client name sent to Excel. Status: All done."))

    def simulate_button_press(self, button, pressed_color="#005BB5", duration=500):
        try:
            original_style = button.styleSheet()
            button.setStyleSheet(original_style + f"\nbackground-color: {pressed_color};")
            QtCore.QTimer.singleShot(duration, lambda: button.setStyleSheet(original_style))
        except Exception as e:
            print(f"simulate_button_press error: {e}")

    def automate_aaa(self):
        self.status_label.setText(self.tr("Status: Starting automation sequence..."))
        self.simulate_button_press(self.process_button, "#005BB5")
        self.delayed_process_input()
        # Adjust timer intervals to allow processing to complete.
        self.aaa_timer1.start(5000)  # Start save after 5 seconds
        self.aaa_timer1.timeout.connect(lambda: (
            self.simulate_button_press(self.save_pdf_json_button, "#005BB5"),
            self.save_data(),
            self.status_label.setText(self.tr("Status: Report created."))
        ))
        self.aaa_timer2.start(8000)  # Start Excel append after 8 seconds
        self.aaa_timer2.timeout.connect(lambda: (
            self.simulate_button_press(self.append_excel_button, "#005BB5"),
            self.append_client_name_to_excel(),
            self.status_label.setText(self.tr("Status: All done.")),
            self.switchToAppointments.emit(self.current_data.get("Name", "Unknown"))
        ))

    def toggle_voice_input(self):
        if self.voice_widget_ar.isEnabled():
            self.voice_widget_ar.setEnabled(False)
            self.pause_voice_btn.setText(self.tr("Resume Voice"))
            self.status_label.setText(self.tr("Status: Voice input paused."))
        else:
            self.voice_widget_ar.setEnabled(True)
            self.pause_voice_btn.setText(self.tr("Pause Voice"))
            self.status_label.setText(self.tr("Status: Voice input resumed."))

    def retranslateUi(self):
        print("ExtractionTab retranslateUi called")
        self.pause_voice_btn.setText(self.tr("Pause Voice"))
        self.header_label.setText(self.tr("Data Extraction"))
        self.prompt_label.setText(self.tr("Enter patient information and AI will analyze it:"))
        self.load_test_button.setText(self.tr("Load Test Data"))
        self.process_button.setText(self.tr("Process Input"))
        self.tableWidget.setHorizontalHeaderLabels([self.tr("Field"), self.tr("Value")])
        self.save_pdf_json_button.setText(self.tr("Create Report"))
        self.append_excel_button.setText(self.tr("Append Name to Clients File"))
        self.start_aaa_btn.setText(self.tr("Start AAA"))
        self.status_label.setText(self.tr("Status: Ready"))
        self.voice_widget_ar.retranslateUi()

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    widget = ExtractionTab()
    widget.show()
    sys.exit(app.exec_())
