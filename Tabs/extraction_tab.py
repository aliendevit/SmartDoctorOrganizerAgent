# extraction_tab.py — Glass-matched Clinical Extraction (LLM-upgraded)
import os
from pathlib import Path
from transformers import AutoTokenizer, AutoModelForCausalLM

HF_CACHE = os.getenv("HF_HOME") or os.path.join(os.path.expanduser("~"), ".cache", "huggingface")
REPO_ID_DEFAULT = "google/gemma-2b-it"  # only used if we *allow* online

def _resolve_local_snapshot():
    """
    Prefer a fully local snapshot from MEDICALDOC_LOCAL_MODEL (same env your chat tab uses),
    which should point to the snapshot folder containing config.json + *.safetensors.
    """
    snap = (os.getenv("MEDICALDOC_LOCAL_MODEL") or "").strip()
    if snap and Path(snap).exists() and (Path(snap) / "config.json").exists():
        return snap
    return ""

import os
import re
import json
import tempfile
from datetime import datetime, timedelta
from typing import Callable, Dict, List, Tuple, Optional
from nlp.local_gemma_it import extract_fields
from PyQt5 import QtWidgets, QtCore, QtGui
try:
    from nlp.local_gemma_it import extract_fields as _gemma_extract
except Exception:
    _gemma_extract = None


__all__ = ["ExtractionTab"]
#---transformer model
try:
    from nlp.transformers_extractor import extract_fields as _hf_extract
except Exception:
    _hf_extract = None
# ---------- Global design tokens (safe fallback if design_system not present)
try:
    from UI.design_system import COLORS as DS_COLORS
except Exception:
    DS_COLORS = {
        "text": "#1f2937", "textDim": "#334155", "muted": "#64748b",
        "primary": "#3A8DFF", "info": "#2CBBA6", "success": "#7A77FF",
        "stroke": "#E5EFFA", "panel": "rgba(255,255,255,0.55)",
        "panelInner": "rgba(255,255,255,0.65)", "inputBg": "rgba(255,255,255,0.88)",
        "stripe": "rgba(240,247,255,0.65)", "selBg": "#3A8DFF", "selFg": "#ffffff",
    }

# ---------- Optional Local LLM (Gemma) extractor ----------
# If available, this provides: extract_structured(text) -> dict with the exact keys we need.
_LLME = None
try:
    from core.ai_assistant import extract_structured as _llm_extract
    _LLME = _llm_extract
except Exception:
    _LLME = None

REPORTLAB_OK = True
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
except Exception:
    REPORTLAB_OK = False
    letter = SimpleDocTemplate = Paragraph = Spacer = Table = TableStyle = None
    colors = getSampleStyleSheet = None

import speech_recognition as sr

# Try to import reports_dir; if missing, fall back to Desktop/reports
try:
    from utils.app_paths import reports_dir
    def _reports_dir() -> str:
        path = reports_dir()
        os.makedirs(path, exist_ok=True)
        return path
except Exception:
    def _reports_dir() -> str:
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        path = os.path.join(desktop, "reports"); os.makedirs(path, exist_ok=True)
        return path

# ---------- Optional SmartExtractor (kept as legacy fallback) ----------
_EXTRACTOR = None
try:
    from nlp.smart_nlp import SmartExtractor
    _EXTRACTOR = SmartExtractor()
except Exception:
    _EXTRACTOR = None

# ---------- Optional spaCy ----------
try:
    import spacy
    try:
        NLP = spacy.load("en_core_web_sm")
    except Exception:
        NLP = None
except Exception:
    NLP = None

# ---------- Optional faster-whisper ----------
try:
    from faster_whisper import WhisperModel
    WHISPER_OK = True
except Exception:
    WhisperModel = None
    WHISPER_OK = False

# ---------- i18n helper ----------
def _tr(obj, text: str) -> str:
    try:
        from features.translation_helper import tr
        return tr(text)
    except Exception:
        return text

# ---------- small style helper ----------
def _polish(*widgets):
    for w in widgets:
        try:
            w.style().unpolish(w); w.style().polish(w); w.update()
        except Exception:
            pass
print(
    "[Extraction] Engines available -> "
    f"LLME(core.ai_assistant)={bool(_LLME)}, "
    f"Gemma(local_gemma_it)={bool(_gemma_extract)}, "
    f"SmartExtractor={bool(_EXTRACTOR)}, "
    f"Transformers={bool(_hf_extract)}"
)

# ====================== Parsing helpers ======================
_DATE_RX = re.compile(r'(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})')
_TIME_RX = re.compile(r'(\d{1,2}:\d{2}\s*[APMapm]{2}|\d{1,2}:\d{2})')
_AR_CHARS = re.compile(r'[\u0600-\u06FF]')

def _today_str() -> str:
    return QtCore.QDate.currentDate().toString("dd-MM-yyyy")

def _safe_dt_parse(date_str: str, fmt_list=("%d-%m-%Y","%d/%m/%Y","%d-%m-%y","%d/%m/%y","%Y-%m-%d")) -> str:
    s = (date_str or "").strip()
    for fmt in fmt_list:
        try:
            return datetime.strptime(s, fmt).strftime("%d-%m-%Y")
        except Exception:
            pass
    # fallback: pick first dd-mm-yyyy-like substring
    m = _DATE_RX.search(s)
    if m:
        return _safe_dt_parse(m.group(1))
    return _today_str()

def _norm_time(s: str) -> str:
    s = (s or "").strip()
    for fmt in ("%I:%M %p","%I:%M%p","%H:%M"):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.strftime("%I:%M %p")
        except Exception:
            pass
    # try to detect inside a longer string
    m = _TIME_RX.search(s)
    if m:
        return _norm_time(m.group(1))
    return "12:00 PM"
# --- NORMALIZATION FOR MODEL OUTPUT ---
# ---------- Key normalization helpers ----------
_NORM_RX = re.compile(r"[\s_\-\.\[\]\(\):]+")

def _nk(s) -> str:
    """Normalize a key: lowercase, remove spaces/underscores/punct for robust matching."""
    return _NORM_RX.sub("", str(s or "")).lower()

def _kv_flat(obj, out=None):
    """
    Flatten nested dict/list structure into a flat key->value map by key name only.
    Last write wins. Also recurses into lists/dicts to surface nested sections like HPI, cc, etc.
    """
    if out is None: out = {}
    if isinstance(obj, dict):
        for k, v in obj.items():
            out[_nk(k)] = v
            _kv_flat(v, out)
    elif isinstance(obj, list):
        for v in obj:
            _kv_flat(v, out)
    return out

def _first_in(flat: dict, keys: list):
    """Return first non-empty value for any alias; if not found, suffix-match."""
    for k in keys:
        v = flat.get(_nk(k))
        if v not in (None, "", [], {}):
            return v
    # suffix fallback: any key that ends with the alias (helps with dotted paths in some models)
    for k, v in flat.items():
        if any(k.endswith(_nk(alias)) for alias in keys):
            if v not in (None, "", [], {}):
                return v
    return None

def _to_listlike(v):
    """Turn model output into a clean list of symptoms."""
    if isinstance(v, list):
        return [str(x).strip() for x in v if str(x).strip()]
    if isinstance(v, dict):
        return [str(x).strip() for x in v.values() if str(x).strip()]
    if isinstance(v, str):
        s = v.strip()
        # Try JSON first
        try:
            j = json.loads(s)
            if isinstance(j, list):
                return [str(x).strip() for x in j if str(x).strip()]
        except Exception:
            pass
        # Split on common separators (English + Arabic " و ")
        parts = re.split(r",|;|/|\\|\band\b| و ", s, flags=re.I)
        return [p.strip() for p in parts if p.strip()]
    return []
def _post_normalize_llm(obj) -> dict:
    flat = _kv_flat(obj)
    out = {
        "Name":            ( _first_in(flat, ["name","patient","patientname"]) or "" ).strip(),
        "Age":             re.search(r"\d{1,3}", str(_first_in(flat, ["age"]) or "" )).group(0)
                           if re.search(r"\d{1,3}", str(_first_in(flat, ["age"]) or "" )) else "",
        "Symptoms":        _to_listlike( _first_in(flat, ["symptoms","chief complaint","cc","presenting complaint"]) or [] ),
        "Notes":           ( _first_in(flat, ["notes","assessment","plan","impression","summary"]) or "" ).strip(),
        "General Date":    str(_first_in(flat, ["general date","date","visit date"]) or ""),
        "Appointment Date":str(_first_in(flat, ["appointment date","appt date"]) or ""),
        "Appointment Time":str(_first_in(flat, ["appointment time","appt time","time"]) or ""),
        "Follow-Up Date":  str(_first_in(flat, ["follow up date","follow-up date","fu date"]) or ""),
    }
    return out

# --- RULE-BASED FALLBACK PARSER ---
def _fallback_parse_patient_info(text: str) -> Dict:
    t = (text or "").strip()

    # Name
    name = ""
    m = re.search(r"(?:^|\b)(?:patient(?: name)?|name)\s*[:\-]\s*([A-Z][a-z]+(?:\s+[A-Z][a-z]+){0,3})", t)
    if not m: m = re.search(r"\b(?:Patient|Pt\.?)\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+){0,3})", t)
    if m: name = m.group(1).strip()

    # Age
    age = ""
    m = re.search(r"\bage\s*[:\-]?\s*(\d{1,3})\b", t, re.I)
    if not m: m = re.search(r"\b(\d{1,3})\s*(?:years?\s*old|yo)\b", t, re.I)
    if m: age = m.group(1)

    # Symptoms
    symptoms: List[str] = []
    sym_rx_list = [
        r"\b(?:presents?|presenting)\s+with\s+(?P<s>.+?)(?:[.;\n]|$)",
        r"\bcc\s*[:\-]\s*(?P<s>.+?)(?:[.;\n]|$)",  # CC: ...
        r"\bchief\s+complaints?\s*[:\-]?\s*(?P<s>.+?)(?:[.;\n]|$)",
        r"\b(?:complain(?:s)?\s+of|c/o)\s+(?P<s>.+?)(?:[.;\n]|$)",
        r"\b(?:reports?|has|experiencing|suffers?\s+from|hx\s*of|h/o)\s+(?P<s>.+?)(?:[.;\n]|$)",
        r"\bsymptoms?\s*[:\-]?\s*(?P<s>.+?)(?:[.;\n]|$)",
    ]
    for rx in sym_rx_list:
        m = re.search(rx, t, re.I)
        if m:
            symptoms = _to_listlike(m.group("s"))
            if symptoms:
                break

    # Notes (English/Arabic)
    notes = ""
    m = re.search(r"\b(?:notes?|assessment|plan|impression)\s*[:\-]\s*(.+?)(?:\n\n|\Z)", t, re.I | re.S)
    if not m:
        m = re.search(r"(?:ملاحظات|التقييم|الخطة)\s*[:\-]\s*(.+?)(?:\n\n|\Z)", t, re.S)
    if m:
        notes = m.group(1).strip()

    # Dates/Times
    appt_date = ""
    appt_time = ""
    m = re.search(r"\bappointment\b.*?"+_DATE_RX.pattern, t, re.I)
    if m:
        m2 = _DATE_RX.search(m.group(0))
        if m2: appt_date = m2.group(1)
    if not appt_date:
        m = _DATE_RX.search(t)
        if m: appt_date = m.group(1)

    m = re.search(r"\bappointment\b.*?"+_TIME_RX.pattern, t, re.I)
    if m:
        m2 = _TIME_RX.search(m.group(0))
        if m2: appt_time = m2.group(1)
    if not appt_time:
        m = _TIME_RX.search(t)
        if m: appt_time = m.group(1)

    follow_up = ""
    m = re.search(r"\bfollow[\-\s]?up\b.*?"+_DATE_RX.pattern, t, re.I)
    if m:
        m2 = _DATE_RX.search(m.group(0))
        if m2: follow_up = m2.group(1)

    gen = _DATE_RX.search(t)
    general_date = gen.group(1) if gen else ""

    return {
        "Name": name or "Unknown",
        "Age": age,
        "Symptoms": symptoms,
        "Notes": notes,
        "General Date": _safe_dt_parse(general_date) if general_date else _today_str(),
        "Appointment Date": _safe_dt_parse(appt_date) if appt_date else "Not Specified",
        "Appointment Time": _norm_time(appt_time) if appt_time else "Not Specified",
        "Follow-Up Date": _safe_dt_parse(follow_up) if follow_up else "Not Specified",
        "Date": _safe_dt_parse(general_date) if general_date else _today_str(),
    }
# --- CANONICAL KEYS USED BY THE APP ---
_CANON = ["Name","Age","Symptoms","Notes","General Date","Appointment Date","Appointment Time","Follow-Up Date","Date"]

def _is_empty(v):
    if v is None: return True
    if isinstance(v, str): return v.strip() == "" or v.strip().lower() in {"n/a","na","none","not specified"}
    if isinstance(v, list): return len(v) == 0
    return False
# put this near _CANON/_is_empty helpers
_ENABLE_PROVENANCE = True

def _track_fill(prev: dict, new: dict, label: str, prov: dict):
    for k in _CANON:
        if _is_empty(prev.get(k)) and not _is_empty(new.get(k)):
            prov[k] = label

def _as_list(v):
    if isinstance(v, list): return v
    if isinstance(v, str):
        parts = re.split(r",|;| and ", v)
        return [p.strip() for p in parts if p.strip()]
    return []

def _merge_extractions(primary: Dict, extra: Dict) -> Dict:
    """Prefer primary; fill blanks from extra. Union lists; normalize dates/times."""
    out = dict(primary or {})

    def pick(k):
        a, b = out.get(k), extra.get(k)
        if k == "Symptoms":
            sa = _as_list(a); sb = _as_list(b)
            out[k] = list(dict.fromkeys([*sa, *sb])) if sa or sb else []
            return
        # numeric age
        if k == "Age":
            if _is_empty(a) and not _is_empty(b): out[k] = re.search(r"\d{1,3}", str(b)).group(0) if re.search(r"\d{1,3}", str(b)) else ""
            return
        # dates and time normalized
        if k in ("General Date","Date","Appointment Date","Follow-Up Date"):
            if _is_empty(a) and not _is_empty(b): out[k] = _safe_dt_parse(str(b))
            return
        if k == "Appointment Time":
            if _is_empty(a) and not _is_empty(b): out[k] = _norm_time(str(b))
            return
        # everything else (Name/Notes)
        if _is_empty(a) and not _is_empty(b): out[k] = b

    for k in _CANON:
        pick(k)

    # Ensure Date mirrors General Date for back-compat
    if _is_empty(out.get("Date")) and not _is_empty(out.get("General Date")):
        out["Date"] = out["General Date"]
    return out

# --- WRAPPERS FOR EACH EXTRACTOR, ALL RETURN CANONICAL SCHEMA ---
def _extract_with_gemma(text: str) -> Dict:
    # 1) core.ai_assistant first
    try:
        if _LLME:
            d = _LLME(text) or {}
            print("[Extraction] LLME returned keys:", list(d.keys()))
            if d: return _post_normalize_llm(d)
    except Exception as e:
        import traceback; print("[Extraction] LLME failed:", e); traceback.print_exc()

    # 2) local_gemma_it
    try:
        if _gemma_extract:
            d = _gemma_extract(text) or {}
            print("[Extraction] Gemma returned keys:", list(d.keys()))
            if d: return _post_normalize_llm(d)
    except Exception as e:
        import traceback; print("[Extraction] Gemma(local_gemma_it) failed:", e); traceback.print_exc()

    return {}

def _extract_with_smart(text: str) -> Dict:
    try:
        if _EXTRACTOR:
            d = _EXTRACTOR.extract(text) or {}
            print("[Extraction] SmartExtractor returned keys:", list(d.keys()))
            if d: return _post_normalize_llm(d)
    except Exception as e:
        import traceback; print("[Extraction] SmartExtractor failed:", e); traceback.print_exc()
    return {}

def _extract_with_transformers(text: str) -> Dict:
    try:
        if _hf_extract:
            d = _hf_extract(text) or {}
            print("[Extraction] Transformers returned keys:", list(d.keys()))
            if d: return _post_normalize_llm(d)
    except Exception as e:
        import traceback; print("[Extraction] Transformers failed:", e); traceback.print_exc()
    return {}


def parse_patient_info(text: str) -> Dict:
    prov = {k: "—" for k in _CANON}

    g = _extract_with_gemma(text)
    merged = dict(g or {})
    _track_fill({}, merged, "gemma", prov)

    s = _extract_with_smart(text)
    m2 = _merge_extractions(merged, s)
    _track_fill(merged, m2, "smart", prov)
    merged = m2

    h = _extract_with_transformers(text)
    m3 = _merge_extractions(merged, h)
    _track_fill(merged, m3, "transformers", prov)
    merged = m3

    rb = _fallback_parse_patient_info(text)  # always run; safe filler
    m4 = _merge_extractions(merged, rb)
    _track_fill(merged, m4, "regex", prov)
    merged = m4

    # guarantee keys
    for k in _CANON:
        merged.setdefault(k, [] if k == "Symptoms" else "")
    if _is_empty(merged.get("General Date")):
        merged["General Date"] = _today_str()
    if _is_empty(merged.get("Date")):
        merged["Date"] = merged["General Date"]

    if _ENABLE_PROVENANCE:
        # console: see where each field came from
        print("Extraction provenance:", prov)
        # optional: expose for UI debugging without breaking tables
        merged["_prov"] = prov  # comment out if you don't want it in JSON exports

    return merged



# ====================== Whisper helpers ======================
USE_WHISPER_BY_DEFAULT = True

def _make_whisper_model(size: str):
    if not WHISPER_OK:
        raise RuntimeError("faster-whisper not installed")
    attempts = [
        ("cuda", "float16"),
        ("auto", "float16"),
        ("cpu",  "int8"),
        ("cpu",  "int8_float32"),
        ("cpu",  "float32"),
    ]
    last_err = None
    for device, ctype in attempts:
        try:
            return WhisperModel(size, device=device, compute_type=ctype)
        except Exception as e:
            last_err = e
    raise RuntimeError(f"Whisper init failed. Last error: {last_err}")

def _lang_to_codes(choice: str):
    if choice == "ar": return "ar", "ar-SA"
    if choice == "en": return "en", "en-US"
    return None, None

class _WhisperThread(QtCore.QThread):
    result = QtCore.pyqtSignal(str)
    error  = QtCore.pyqtSignal(str)
    def __init__(self, wav_bytes: bytes, language=None, model_size="base", translate=False):
        super().__init__()
        self.wav_bytes = wav_bytes
        self.language = language
        self.model_size = model_size
        self.translate = bool(translate)

    def run(self):
        tmp_path = None
        try:
            fd, tmp_path = tempfile.mkstemp(suffix=".wav")
            os.write(fd, self.wav_bytes); os.close(fd)
            if not hasattr(_WhisperThread, "_model"):
                _WhisperThread._model = _make_whisper_model(self.model_size)
            model = _WhisperThread._model

            segments, _info = model.transcribe(
                tmp_path,
                language=self.language,
                vad_filter=True,
                beam_size=5,
                condition_on_previous_text=False,
                task=("translate" if self.translate else "transcribe")
            )
            text = " ".join(seg.text.strip() for seg in segments).strip()
            self.result.emit(text)
        except Exception as e:
            self.error.emit(str(e))
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try: os.remove(tmp_path)
                except Exception: pass

# ====================== Voice Input (no pause control) ======================
class VoiceInputWidget(QtWidgets.QWidget):
    textReady = QtCore.pyqtSignal(str)

    def __init__(self, parent=None, language="en-US", use_whisper=None, whisper_model_size="base"):
        super().__init__(parent)
        lang_l = (language or "").lower()
        if lang_l.startswith("ar"): self.choice = "ar"
        elif lang_l.startswith("en"): self.choice = "en"
        else: self.choice = "auto"

        self.use_whisper = USE_WHISPER_BY_DEFAULT if use_whisper is None else bool(use_whisper)
        self.whisper_model_size = whisper_model_size
        self._build_ui()
        self._refresh_labels()

    def _build_ui(self):
        root = QtWidgets.QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0); root.setSpacing(8)

        row = QtWidgets.QHBoxLayout(); row.setSpacing(8)
        self.lbl = QtWidgets.QLabel()
        self.combo = QtWidgets.QComboBox()
        self.combo.addItem("Auto", "auto")
        self.combo.addItem("Arabic (ar)", "ar")
        self.combo.addItem("English (en)", "en")
        idx = max(0, self.combo.findData(self.choice))
        self.combo.setCurrentIndex(idx)
        self.combo.currentIndexChanged.connect(self._on_lang_change)

        self.chk_translate = QtWidgets.QCheckBox()
        self.chk_translate.setToolTip("Whisper can translate Arabic → English when enabled.")
        self.chk_translate.stateChanged.connect(self._refresh_labels)

        for w in (self.combo, self.chk_translate):
            w.setProperty("variant", "ghost"); w.setProperty("accent", "slate")
        _polish(self.combo, self.chk_translate)

        row.addWidget(self.lbl)
        row.addWidget(self.combo, 1)
        row.addWidget(self.chk_translate)
        root.addLayout(row)

        self.btn = QtWidgets.QPushButton()
        self.btn.setMinimumHeight(44)
        self.btn.setProperty("variant", "info"); self.btn.setProperty("accent", "sky")
        self.btn.clicked.connect(self._start)
        _polish(self.btn)
        root.addWidget(self.btn)

    def _on_lang_change(self):
        self.choice = self.combo.currentData() or "auto"
        self._refresh_labels()

    def _refresh_labels(self):
        self.lbl.setText(_tr(self, "ASR Language:"))
        self.chk_translate.setText(_tr(self, "Translate to English (Whisper)"))
        engine = "Whisper" if (self.use_whisper and WHISPER_OK) else "Google"
        c = {"auto": _tr(self, "Auto"), "ar": "ar", "en": "en"}[self.choice]
        extra = _tr(self, " · translate") if (engine == "Whisper" and self.chk_translate.isChecked()) else ""
        self.btn.setText(f"{_tr(self,'Voice Input')} ({c}) · {engine}{extra}")

    def retranslateUi(self):
        self._refresh_labels()

    def _start(self):
        self.btn.setDown(True); QtCore.QTimer.singleShot(150, lambda: self.btn.setDown(False))
        self.btn.setText(_tr(self, "Listening…")); QtWidgets.QApplication.processEvents()
        r = sr.Recognizer()
        try:
            with sr.Microphone() as source:
                r.adjust_for_ambient_noise(source, duration=0.5)
                audio = r.listen(source, timeout=10)
        except sr.WaitTimeoutError:
            QtWidgets.QMessageBox.warning(self, _tr(self,"Voice Input"), _tr(self,"Listening timed out."))
            self._refresh_labels(); return
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, _tr(self,"Voice Input"), f"{_tr(self,'Microphone error:')} {e}")
            self._refresh_labels(); return

        w_lang, g_lang = _lang_to_codes(self.choice)

        if self.use_whisper and WHISPER_OK:
            self.btn.setText(_tr(self, "Transcribing… (Whisper)")); QtWidgets.QApplication.processEvents()
            t = _WhisperThread(audio.get_wav_data(), language=w_lang,
                               model_size=self.whisper_model_size,
                               translate=self.chk_translate.isChecked())
            t.result.connect(self._ok); t.error.connect(self._err)
            self._t = t; t.start()
            return

        self.btn.setText(_tr(self, "Transcribing… (Google)")); QtWidgets.QApplication.processEvents()
        try:
            if g_lang is None:
                text = self._google_dual(r, audio)
            else:
                text = r.recognize_google(audio, language=g_lang)
            self._ok(text)
        except sr.UnknownValueError:
            self._err(_tr(self, "Could not understand the audio."))
        except sr.RequestError as e:
            self._err(f"{_tr(self,'Speech service error:')} {e}")
        except Exception as e:
            self._err(str(e))

    def _google_dual(self, r: sr.Recognizer, audio: sr.AudioData) -> str:
        cands = []
        for code in ("en-US", "ar-SA"):
            try:
                t = r.recognize_google(audio, language=code).strip()
                if t: cands.append((code, t))
            except Exception:
                pass
        if not cands:
            raise sr.RequestError("Google ASR returned no candidates")
        def ar_ratio(s: str): return len(_AR_CHARS.findall(s))/max(1,len(s))
        best = max(cands, key=lambda kv: (ar_ratio(kv[1]) if kv[0].startswith("ar") else 0.0, len(kv[1])) )
        return best[1]

    def _ok(self, text: str):
        self.textReady.emit(text or "")
        self._refresh_labels()

    def _err(self, msg: str):
        QtWidgets.QMessageBox.warning(
            self, _tr(self,"Voice Input Error"),
            msg + "\n" + _tr(self,"If Whisper is unavailable, the tool will fall back to Google when possible.")
        )
        self._refresh_labels()

# ====================== Report helpers ======================
def generate_pdf_report(data: Dict, file_path: str):
    if not REPORTLAB_OK:
        raise RuntimeError('ReportLab is not installed. Install `reportlab` to enable PDF export.')
    doc = SimpleDocTemplate(file_path, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []
    nm = data.get("Name","Unknown")
    elements.append(Paragraph(f"Patient Report: {nm}", styles["Title"]))
    elements.append(Spacer(1, 12))

    # Optional brief notes summary at top if provided
    summary = (data.get("Notes") or "").strip()
    if summary:
        elements.append(Paragraph(f"<b>Notes:</b><br/>{summary}", styles["BodyText"]))
        elements.append(Spacer(1, 12))

    header = [Paragraph("<b>Field</b>", styles["BodyText"]), Paragraph("<b>Value</b>", styles["BodyText"])]
    rows = [header]
    fields = [
        ("Age", data.get("Age","N/A")),
        ("Symptoms", ", ".join(data.get("Symptoms",[]))),
        ("General Date", data.get("General Date","Not Specified")),
        ("Appointment Date", data.get("Appointment Date","Not Specified")),
        ("Appointment Time", data.get("Appointment Time","Not Specified")),
        ("Follow-Up Date", data.get("Follow-Up Date","Not Specified")),
    ]
    for k,v in fields:
        rows.append([Paragraph(k, styles["BodyText"]), Paragraph(str(v), styles["BodyText"])])

    table = Table(rows, colWidths=[150, 350])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4f46e5')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN',      (0,0), (-1,-1), 'LEFT'),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#f3f4f6')),
        ('GRID',       (0,0), (-1,-1), 0.5, colors.HexColor('#e5e7eb')),
    ]))
    elements.append(table)
    doc.build(elements)

# ====================== Minimal Agent & actions ======================
class Agent(QtCore.QObject):
    log = QtCore.pyqtSignal(str)
    step_started = QtCore.pyqtSignal(str)
    step_line = QtCore.pyqtSignal(str)
    step_finished = QtCore.pyqtSignal(str, str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._actions: Dict[str, Callable[[Dict], Tuple[Dict, List[str]]]] = {}

    def register(self, name: str, fn: Callable[[Dict], Tuple[Dict, List[str]]]):
        self._actions[name] = fn

    def run_plan(self, steps: List[str], ctx: Dict) -> Dict:
        self.log.emit("Agent: starting plan")
        for name in steps:
            self.step_started.emit(name)
            fn = self._actions.get(name)
            if not fn:
                self.step_line.emit(f"⚠️ step '{name}' not found; skipping")
                self.step_finished.emit(name, "skipped")
                continue
            new_ctx, lines = fn(dict(ctx))
            for ln in lines:
                self.step_line.emit(ln)
            ctx.update(new_ctx or {})
            self.step_finished.emit(name, "ok")
        self.log.emit("Agent: plan complete")
        return ctx

def action_insert_db(ctx: Dict) -> Tuple[Dict, List[str]]:
    lines = ["Writing patient to database…"]
    try:
        from data.data import insert_client
        insert_client(ctx.get("data", {}))
        lines.append("✅ Inserted.")
    except Exception:
        lines.append("ℹ️ data.data.insert_client not available; simulated.")
    return ctx, lines

def action_followup_rule(ctx: Dict) -> Tuple[Dict, List[str]]:
    d = dict(ctx.get("data", {}))
    lines = ["Applying follow-up rule…"]
    fu = (d.get("Follow-Up Date") or "").strip()
    if not fu or fu == "Not Specified":
        base_str = _safe_dt_parse(d.get("General Date") or d.get("Date") or _today_str())
        base = datetime.strptime(base_str, "%d-%m-%Y")
        d["Follow-Up Date"] = (base + timedelta(days=7)).strftime("%d-%m-%Y")
        lines.append(f"Set follow-up to {d['Follow-Up Date']}.")
    else:
        lines.append("Follow-up already present.")
    ctx["data"] = d
    return ctx, lines

def action_tag_status(ctx: Dict) -> Tuple[Dict, List[str]]:
    d = dict(ctx.get("data", {}))
    lines = ["Tagging appointment status…"]
    apd = d.get("Appointment Date","Not Specified")
    apt = d.get("Appointment Time","Not Specified")
    d["Appointment Status"] = "Scheduled" if (apd != "Not Specified" and apt != "Not Specified") else "Missing"
    lines.append(f"Status: {d['Appointment Status']}")
    ctx["data"] = d
    return ctx, lines

def action_generate_pdf(ctx: Dict) -> Tuple[Dict, List[str]]:
    d = dict(ctx.get("data", {}))
    nm = d.get("Name","Unknown")
    safe = "".join(c for c in nm if c.isalnum() or c in (" ","_")).replace(" ","_") or "Unknown"
    pdf = os.path.join(_reports_dir(), f"{safe}_report.pdf")
    lines = ["Generating PDF report…"]
    try:
        generate_pdf_report(d, pdf)
        ctx["pdf_path"] = pdf
        lines.append(f"✅ PDF: {pdf}")
    except Exception as e:
        lines.append(f"❌ PDF failed: {e}")
    return ctx, lines

def action_write_json(ctx: Dict) -> Tuple[Dict, List[str]]:
    d = dict(ctx.get("data", {}))
    nm = d.get("Name","Unknown")
    safe = "".join(c for c in nm if c.isalnum() or c in (" ","_")).replace(" ","_") or "Unknown"
    jsn = os.path.join(_reports_dir(), f"{safe}_report.json")
    lines = ["Writing JSON…"]
    try:
        with open(jsn, "w", encoding="utf-8") as f:
            json.dump(d, f, indent=4, ensure_ascii=False)
        ctx["json_path"] = jsn
        lines.append(f"✅ JSON: {jsn}")
    except Exception as e:
        lines.append(f"❌ JSON failed: {e}")
    return ctx, lines

# ====================== Agent Simulator Dialog ======================
class AgentRunDialog(QtWidgets.QDialog):
    """
    Real-time agent run:
    - Uses your Agent's real actions (insert_db, followup_rule, etc.)
    - Streams step_started / step_line / step_finished / log into the UI
    - Runs in a worker thread; no freezing, no fake delays
    """
    def __init__(self, agent: Agent, steps: List[str], ctx: Dict, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Agent (live)")
        self.setAttribute(QtCore.Qt.WA_DeleteOnClose, True)
        self.resize(780, 560)

        self.agent = agent
        self.steps = list(steps)
        self.ctx = dict(ctx or {})

        v = QtWidgets.QVBoxLayout(self)

        # Header row
        head = QtWidgets.QHBoxLayout()
        self.title = QtWidgets.QLabel("Running plan")
        self.title.setStyleSheet("font-size:16px; font-weight:700;")
        head.addWidget(self.title); head.addStretch(1)
        self.btn_run = QtWidgets.QPushButton("Run")
        self.btn_close = QtWidgets.QPushButton("Close")
        head.addWidget(self.btn_run); head.addWidget(self.btn_close)
        v.addLayout(head)

        # Live log
        self.log = QtWidgets.QTextEdit(readOnly=True)
        self.log.setStyleSheet("font-family: Consolas, Menlo, monospace;")
        v.addWidget(self.log, 1)

        # Files row (for PDF/JSON buttons the actions may create)
        self.files_row = QtWidgets.QHBoxLayout()
        self.files_row.addStretch(1)
        v.addLayout(self.files_row)

        # Wire buttons
        self.btn_close.clicked.connect(self.close)
        self.btn_run.clicked.connect(self._run)

        # Stream agent signals to the dialog
        self.agent.log.connect(self._append)
        self.agent.step_started.connect(lambda n: self._append(f"\n▶️  {n}"))
        self.agent.step_line.connect(lambda s: self._append(f"   {s}"))
        self.agent.step_finished.connect(lambda n, r: self._append(f"✅ done: {n}"))

        self.worker = None

    def _append(self, line: str):
        self.log.append(line)
        self.log.ensureCursorVisible()

    def _add_file_button(self, label: str, path: str):
        btn = QtWidgets.QPushButton(label)
        btn.clicked.connect(lambda: QtGui.QDesktopServices.openUrl(QtCore.QUrl.fromLocalFile(path)))
        self.files_row.insertWidget(self.files_row.count()-1, btn)

    def _run(self):
        if self.worker:
            return
        self.btn_run.setEnabled(False)
        self._append("Starting…")

        # Use your existing _AgentWorker which calls agent.run_plan()
        self.worker = _AgentWorker(self.agent, list(self.steps), dict(self.ctx))
        self.worker.done.connect(self._done)
        self.worker.failed.connect(self._fail)
        self.worker.start()

    def _done(self, ctx_out: Dict):
        # If actions generated files, add open buttons
        pdf = ctx_out.get("pdf_path"); jsn = ctx_out.get("json_path")
        if pdf: self._add_file_button("Open PDF", pdf)
        if jsn: self._add_file_button("Open JSON", jsn)
        self._append("\n🎉 PLAN COMPLETE")
        self.btn_run.setEnabled(True)
        self.worker = None

    def _fail(self, err: str):
        self._append(f"\n❌ FAILED: {err}")
        QtWidgets.QMessageBox.critical(self, "Agent", err)
        self.btn_run.setEnabled(True)
        self.worker = None


class _AgentWorker(QtCore.QThread):
    done = QtCore.pyqtSignal(dict)
    failed = QtCore.pyqtSignal(str)
    def __init__(self, agent: Agent, steps: List[str], ctx: Dict):
        super().__init__()
        self.agent = agent; self.steps = steps; self.ctx = ctx
    def run(self):
        try:
            out = self.agent.run_plan(self.steps, self.ctx)
            self.done.emit(out)
        except Exception as e:
            self.failed.emit(str(e))

# ====================== Main Extraction Tab ======================
class ExtractionTab(QtWidgets.QWidget):
    dataProcessed = QtCore.pyqtSignal(dict)
    appointmentProcessed = QtCore.pyqtSignal(dict)
    switchToAppointments = QtCore.pyqtSignal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._settings = QtCore.QSettings("YourOrg", "MedicalDocAI Demo v1.9.3")
        self._setup_ui()
        self._build_agent()
        self._restore_state()

    def tr(self, text): return _tr(self, text)

    # ---------- UI ----------
    def _setup_ui(self):
        root = QtWidgets.QVBoxLayout(self); root.setContentsMargins(16,16,16,16); root.setSpacing(12)

        # Header
        header = QtWidgets.QFrame(); header.setProperty("modernCard", True)
        h = QtWidgets.QHBoxLayout(header); h.setContentsMargins(12,12,12,12); h.setSpacing(8)
        title = QtWidgets.QLabel(self.tr("Clinical Extraction"))
        title.setStyleSheet("font: 700 18pt 'Segoe UI';")
        subtitle = QtWidgets.QLabel(self.tr("Dictate or paste—AI structures the visit and fills the report table."))
        subtitle.setStyleSheet(f"color:{DS_COLORS['muted']};")
        left = QtWidgets.QVBoxLayout(); left.addWidget(title); left.addWidget(subtitle)
        h.addLayout(left); h.addStretch(1)
        root.addWidget(header)

        # Split: Left (Input + Voice), Right (Preview)
        split = QtWidgets.QSplitter(QtCore.Qt.Horizontal); split.setChildrenCollapsible(False)
        root.addWidget(split, 1)

        # ----- LEFT: Input -----
        left_card = QtWidgets.QFrame(); left_card.setProperty("modernCard", True)
        lc = QtWidgets.QVBoxLayout(left_card); lc.setContentsMargins(12,12,12,12); lc.setSpacing(8)

        lbl = QtWidgets.QLabel(self.tr("Patient narrative (Arabic/English)."))
        lbl.setStyleSheet(f"color:{DS_COLORS['textDim']};")
        self.txt = QtWidgets.QTextEdit(); self.txt.setMinimumHeight(180)
        self.txt.setPlaceholderText(self.tr("Example: Patient Jane Smith, age 23, complains of cough and headache. Appointment 21-11-2025 at 10:30 AM. Follow-up 28-11-2025."))
        lc.addWidget(lbl); lc.addWidget(self.txt, 1)

        # Voice strip
        voice_strip = QtWidgets.QFrame()
        vs = QtWidgets.QHBoxLayout(voice_strip); vs.setContentsMargins(0,0,0,0); vs.setSpacing(8)
        self.voice = VoiceInputWidget(language="ar-SA", use_whisper=True, whisper_model_size="base")
        self.voice.textReady.connect(lambda s: self.txt.setPlainText(s))
        vs.addWidget(self.voice, 1)
        lc.addWidget(voice_strip)

        # Actions row
        actions = QtWidgets.QHBoxLayout(); actions.setSpacing(8)
        self.btn_process = QtWidgets.QPushButton(self.tr("Process (Ctrl+Enter)"))
        self.btn_process.setProperty("variant","success"); _polish(self.btn_process)
        self.btn_process.clicked.connect(self._delayed_process)

        self.btn_load = QtWidgets.QPushButton(self.tr("Load Sample (Ctrl+L)"))
        self.btn_load.setProperty("variant","ghost"); _polish(self.btn_load)
        self.btn_load.clicked.connect(self._load_test)

        self.btn_agent = QtWidgets.QPushButton(self.tr("Agent (F1)"))
        self.btn_agent.setProperty("variant","ghost"); _polish(self.btn_agent)
        self.btn_agent.clicked.connect(self._open_agent)

        actions.addWidget(self.btn_process); actions.addStretch(1); actions.addWidget(self.btn_load); actions.addWidget(self.btn_agent)
        lc.addLayout(actions)

        split.addWidget(left_card)

        # ----- RIGHT: Preview -----
        right_card = QtWidgets.QFrame(); right_card.setProperty("modernCard", True)
        rc = QtWidgets.QVBoxLayout(right_card); rc.setContentsMargins(12,12,12,12); rc.setSpacing(8)

        self.table = QtWidgets.QTableWidget(0,2)
        self.table.setHorizontalHeaderLabels([self.tr("Field"), self.tr("Value")])
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectRows)
        self.table.setEditTriggers(QtWidgets.QAbstractItemView.NoEditTriggers)
        rc.addWidget(self.table, 1)

        # Export row
        export = QtWidgets.QHBoxLayout(); export.setSpacing(8)
        self.btn_report = QtWidgets.QPushButton(self.tr("Create Report (PDF+JSON)"))
        self.btn_report.setProperty("variant","info"); _polish(self.btn_report)
        self.btn_report.clicked.connect(self._save_report)

        self.btn_excel  = QtWidgets.QPushButton(self.tr("Append Name to Excel"))
        self.btn_excel.setProperty("variant","ghost"); _polish(self.btn_excel)
        self.btn_excel.clicked.connect(self._append_excel)

        export.addWidget(self.btn_report); export.addStretch(1); export.addWidget(self.btn_excel)
        rc.addLayout(export)

        split.addWidget(right_card)
        split.setStretchFactor(0, 1)
        split.setStretchFactor(1, 1)

        # Status bar
        status = QtWidgets.QFrame(); status.setProperty("modernCard", True)
        st = QtWidgets.QHBoxLayout(status); st.setContentsMargins(12,10,12,10); st.setSpacing(8)
        self.lbl_status = QtWidgets.QLabel(self.tr("Status: Ready")); self.lbl_status.setStyleSheet("font-weight:600;")
        st.addWidget(self.lbl_status)
        root.addWidget(status)

        # Shortcuts
        QtWidgets.QShortcut(QtGui.QKeySequence("Ctrl+Return"), self, activated=self._delayed_process)
        QtWidgets.QShortcut(QtGui.QKeySequence("Ctrl+Enter"), self, activated=self._delayed_process)
        QtWidgets.QShortcut(QtGui.QKeySequence("Ctrl+L"), self, activated=self._load_test)
        QtWidgets.QShortcut(QtGui.QKeySequence("F1"), self, activated=self._open_agent)

        # Timers (kept)
        self._t1 = QtCore.QTimer(self); self._t1.setSingleShot(True); self._t1.timeout.connect(self._save_report)
        self._t2 = QtCore.QTimer(self); self._t2.setSingleShot(True); self._t2.timeout.connect(self._append_excel)

        # Apply glass theme for this tab
        self.setStyleSheet(self._tab_qss())
#tab
    def _normalize_appointment(self, data: Dict) -> Dict:
        """Ensure dates/times exist and are formatted for downstream tabs."""

        def _safe_dt_parse(date_str: str, fmt_list=("%d-%m-%Y", "%d/%m/%Y", "%d-%m-%y", "%d/%m/%y", "%Y-%m-%d")) -> str:
            s = (date_str or "").strip()
            for fmt in fmt_list:
                try:
                    return datetime.strptime(s, fmt).strftime("%d-%m-%Y")
                except Exception:
                    pass
            # fall back to today if unparseable
            return QtCore.QDate.currentDate().toString("dd-MM-yyyy")

        def _norm_time(s: str) -> str:
            s = (s or "").strip()
            for fmt in ("%I:%M %p", "%I:%M%p", "%H:%M"):
                try:
                    dt = datetime.strptime(s, fmt)
                    return dt.strftime("%I:%M %p").lstrip("0")
                except Exception:
                    pass
            return "12:00 PM"

        d = dict(data or {})
        d["Date"] = _safe_dt_parse(d.get("Date"))
        ad = d.get("Appointment Date")
        at = d.get("Appointment Time")
        d["Appointment Date"] = _safe_dt_parse(
            ad) if ad and ad != "Not Specified" else QtCore.QDate.currentDate().toString("dd-MM-yyyy")
        d["Appointment Time"] = _norm_time(at) if at and at != "Not Specified" else "12:00 PM"
        return d

    # Glassy QSS for this tab
    def _tab_qss(self) -> str:
        p = DS_COLORS
        return f"""
        QWidget {{ color:{p['text']}; font-family:'Segoe UI', Arial; font-size:14px; }}
        QFrame[modernCard="true"] {{
            background:{p['panel']};
            border:1px solid rgba(255,255,255,0.45);
            border-radius:12px;
        }}
        QLineEdit, QComboBox, QTextEdit {{
            background:{p['inputBg']};
            color:#0f172a;
            border:1px solid #D6E4F5;
            border-radius:8px;
            padding:6px 10px;
            selection-background-color:{p['selBg']};
            selection-color:{p['selFg']};
        }}
        QLineEdit:focus, QComboBox:focus, QTextEdit:focus {{
            border:1px solid {p['primary']};
            box-shadow:0 0 0 2px rgba(58,141,255,0.18);
        }}
        QPushButton {{
            border-radius:10px; padding:9px 14px; font-weight:600;
            border:1px solid transparent; background:{p['primary']}; color:white;
        }}
        QPushButton:hover {{ filter:brightness(1.05); }}
        QPushButton:pressed {{ filter:brightness(0.95); }}
        QPushButton[variant="ghost"] {{
            background: rgba(255,255,255,0.85); color:#0F172A; border:1px solid #D6E4F5;
        }}
        QPushButton[variant="ghost"]:hover {{ background: rgba(255,255,255,0.95); }}
        QPushButton[variant="success"] {{ background:{p['success']}; color:white; }}
        QPushButton[variant="info"]    {{ background:{p['info']};    color:white; }}
        QHeaderView::section {{
            background: rgba(255,255,255,0.85);
            color:#334155;
            padding:8px 10px;
            border:0; border-bottom:1px solid {p['stroke']};
            font-weight:600;
        }}
        QTableWidget, QTableView {{
            background:{p['panelInner']};
            color:#0f172a;
            border:1px solid {p['stroke']};
            border-radius:10px;
            gridline-color:#E8EEF7;
            selection-background-color:{p['selBg']};
            selection-color:{p['selFg']};
        }}
        QTableView::item:!selected:alternate {{ background:{p['stripe']}; }}
        QScrollBar:vertical {{ background:transparent; width:10px; margin:4px; }}
        QScrollBar::handle:vertical {{ background:rgba(58,141,255,0.55); min-height:28px; border-radius:6px; }}
        QScrollBar:horizontal {{ background:transparent; height:10px; margin:4px; }}
        QScrollBar::handle:horizontal {{ background:rgba(58,141,255,0.55); min-width:28px; border-radius:6px; }}
        QScrollBar::add-line, QScrollBar::sub-line {{ width:0; height:0; }}
        """

    # ---------- Agent ----------
    def _build_agent(self):
        self.agent = Agent(self)
        self.agent.register("insert_db", action_insert_db)
        self.agent.register("followup_rule", action_followup_rule)
        self.agent.register("tag_status", action_tag_status)
        self.agent.register("generate_pdf", action_generate_pdf)
        self.agent.register("write_json", action_write_json)
        self.agent.log.connect(lambda s: self.lbl_status.setText(s))

    # ---------- Persistence ----------
    def _restore_state(self):
        try:
            last_text = self._settings.value("extraction/last_text", "", type=str)
            last_lang = self._settings.value("extraction/last_lang", "auto", type=str)
            if last_text:
                self.txt.setPlainText(last_text)
            # Voice widget may not be built when restoring — guard:
            if hasattr(self, "voice") and self.voice and isinstance(self.voice, QtWidgets.QWidget):
                i = self.voice.combo.findData(last_lang)
                if i >= 0:
                    self.voice.combo.setCurrentIndex(i)
        except Exception:
            pass

    def _save_state(self):
        try:
            self._settings.setValue("extraction/last_text", self.txt.toPlainText())
            if hasattr(self, "voice") and self.voice and isinstance(self.voice, QtWidgets.QWidget):
                self._settings.setValue("extraction/last_lang", self.voice.combo.currentData() or "auto")
        except Exception:
            pass

    _tokenizer = None
    _model = None

    def _load():
        global _tokenizer, _model
        if _tokenizer is not None and _model is not None:
            return _tokenizer, _model

        local_snap = _resolve_local_snapshot()
        if local_snap:
            # ✅ Fully offline path (your case)
            _tokenizer = AutoTokenizer.from_pretrained(local_snap, local_files_only=True)
            _model = AutoModelForCausalLM.from_pretrained(local_snap, local_files_only=True)
            return _tokenizer, _model

        # Optional: allow online if your environment ever permits it
        repo_or_path = os.getenv("GEMMA_REPO_ID", REPO_ID_DEFAULT)
        allow_online = os.getenv("ALLOW_HF_ONLINE", "0") == "1"
        _tokenizer = AutoTokenizer.from_pretrained(
            repo_or_path, cache_dir=HF_CACHE, local_files_only=not allow_online
        )
        _model = AutoModelForCausalLM.from_pretrained(
            repo_or_path, cache_dir=HF_CACHE, local_files_only=not allow_online
        )
        return _tokenizer, _model

    # ---------- Actions ----------
    def _load_test(self):
        tx = ("Patient Jane Smith, age 23, complains of cough and headache. "
              "Appointment: 21-11-2025 at 10:30 AM. "
              "Follow-up: 28-11-2025. "
              "Notes: Mild respiratory distress, advice fluids and rest.")
        self.txt.setPlainText(tx)
        self.lbl_status.setText(self.tr("Status: Sample loaded."))

    def _normalize_for_app(self, d: Dict) -> Dict:
        """Ensure required keys exist & are formatted; mirror General Date → Date for app compatibility."""
        out = dict(d or {})
        # Normalize & defaults for the eight required fields
        out["Name"] = (out.get("Name") or "Unknown") or "Unknown"
        out["Age"] = out.get("Age") if (out.get("Age") not in ("", None)) else ""
        out["Symptoms"] = out.get("Symptoms") or []
        out["Notes"] = out.get("Notes") or ""
        out["General Date"] = _safe_dt_parse(out.get("General Date") or out.get("Date") or _today_str())
        out["Appointment Date"] = (_safe_dt_parse(out.get("Appointment Date"))
                                   if out.get("Appointment Date") else "Not Specified")
        out["Appointment Time"] = (_norm_time(out.get("Appointment Time"))
                                   if out.get("Appointment Time") else "Not Specified")
        out["Follow-Up Date"] = (_safe_dt_parse(out.get("Follow-Up Date"))
                                 if out.get("Follow-Up Date") else "Not Specified")
        # Back-compat for other tabs
        out["Date"] = out["General Date"]
        return out

    def _delayed_process(self):
        self.lbl_status.setText(self.tr("Status: Processing input…"))
        self._save_state()
        self._thinking = QtWidgets.QMessageBox(self)
        self._thinking.setWindowTitle(self.tr("Processing"))
        self._thinking.setText(self.tr("Generating output… Please wait."))
        self._thinking.setStandardButtons(QtWidgets.QMessageBox.NoButton)
        self._thinking.show()
        self.txt.setDisabled(True)
        self.btn_process.setDown(True); QtCore.QTimer.singleShot(150, lambda: self.btn_process.setDown(False))
        QtCore.QTimer.singleShot(350, self._process)

    def _process(self):
        try:
            raw = self.txt.toPlainText().strip()
            if not raw:
                QtWidgets.QMessageBox.warning(self, self.tr("Input Error"), self.tr("Please enter dictation or text."))
                return

            extracted = parse_patient_info(raw)
            self.current_data = self._normalize_for_app(extracted)

            appt_payload = dict(self.current_data)  # already normalized

            self._populate_table(self.current_data)
            self.dataProcessed.emit(dict(self.current_data))
            self.appointmentProcessed.emit(dict(appt_payload))
            self.switchToAppointments.emit(appt_payload.get("Name","Unknown"))

            try:
                from data.data import insert_client
                insert_client(self.current_data)
            except Exception:
                pass

            self.lbl_status.setText(self.tr("Status: Input processed successfully."))
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Processing Error"), self.tr("An error occurred:\n") + str(e))
            self.lbl_status.setText(self.tr("Status: Error processing input."))
        finally:
            self.txt.setDisabled(False)
            if hasattr(self, "_thinking"):
                self._thinking.hide()

    def _populate_table(self, data: Dict):
        """Fill the report table with EXACT fields requested."""
        self.table.setRowCount(0)
        fnt = QtGui.QFont("Segoe UI", 11)
        order = [
            "Name", "Age", "Symptoms", "Notes",
            "General Date", "Appointment Date", "Appointment Time", "Follow-Up Date"
        ]
        for key in order:
            val = data.get(key, "")
            if isinstance(val, list):
                val = ", ".join(val)
            row = self.table.rowCount()
            self.table.insertRow(row)
            it1 = QtWidgets.QTableWidgetItem(key); it1.setFont(fnt)
            it2 = QtWidgets.QTableWidgetItem(str(val)); it2.setFont(fnt)
            self.table.setItem(row, 0, it1); self.table.setItem(row, 1, it2)

    def _save_report(self):
        if not getattr(self, "current_data", None):
            QtWidgets.QMessageBox.warning(self, self.tr("Save Data Error"), self.tr("Please process input first."))
            return
        self.lbl_status.setText(self.tr("Status: Saving report…"))
        try:
            nm = self.current_data.get("Name","Unknown")
            safe = "".join(c for c in nm if c.isalnum() or c in (" ","_")).replace(" ","_") or "Unknown"
            pdf = os.path.join(_reports_dir(), f"{safe}_report.pdf")
            jsn = os.path.join(_reports_dir(), f"{safe}_report.json")
            generate_pdf_report(self.current_data, pdf)
            with open(jsn, "w", encoding="utf-8") as f:
                json.dump(self.current_data, f, indent=4, ensure_ascii=False)
            QtWidgets.QMessageBox.information(self, self.tr("Report"), self.tr("Saved:\n") + pdf + "\n" + jsn)
            self.lbl_status.setText(self.tr("Status: Report created."))
        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Save Data Error"), str(e))
            self.lbl_status.setText(self.tr("Status: Error saving report."))

    def _append_excel(self):
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            QtWidgets.QMessageBox.warning(self, self.tr("Excel Error"),
                                          self.tr("openpyxl is required. Install with 'pip install openpyxl'."))
            return
        path = os.path.join(os.path.expanduser("~"), "Desktop", "clients.xlsx")
        if os.path.exists(path):
            wb = load_workbook(path); ws = wb.active
        else:
            wb = Workbook(); ws = wb.active; ws["A1"] = "Client Name"
        nm = (getattr(self, "current_data", {}) or {}).get("Name","Unknown")
        ws.append([nm]); wb.save(path)
        QtWidgets.QMessageBox.information(self, self.tr("Excel"), self.tr("Appended to: ") + path)
        self.lbl_status.setText(self.tr("Status: Client name sent to Excel."))

    def _resolve_compute_mode() -> str:
        mode = str(AS.read_all().get("ai/compute_mode", "auto"))
        if mode == "gpu" and torch.cuda.is_available(): return "cuda"
        if mode == "cpu": return "cpu"
        return "cuda" if torch.cuda.is_available() else "cpu"

    def _make_whisper_model(size: str):
        if not WHISPER_OK:
            raise RuntimeError("faster-whisper not installed")
        device = _resolve_compute_mode()
        attempts = [
            (device, "float16" if device == "cuda" else "int8"),
            ("auto", "float16"),
            ("cpu", "int8"),
            ("cpu", "float32"),
        ]
        last_err = None
        for dev, ctype in attempts:
            try:
                return WhisperModel(size, device=dev, compute_type=ctype)
            except Exception as e:
                last_err = e
        raise RuntimeError(f"Whisper init failed. Last error: {last_err}")
    def _open_agent(self):
        """
        Real-time Agent:
        - Parse input
        - Show live LLM narration as each step runs
        """
        try:
            raw = self.txt.toPlainText().strip()
            if not raw:
                QtWidgets.QMessageBox.warning(self, self.tr("Input Error"), self.tr("Please enter dictation or text."))
                return

            # Parse exactly like Process
            self.current_data = parse_patient_info(raw) or {}
            appt_payload = self._normalize_appointment(self.current_data)

            self._populate_table(self.current_data)
            self.dataProcessed.emit(dict(self.current_data))
            self.appointmentProcessed.emit(dict(appt_payload))
            self.switchToAppointments.emit(appt_payload.get("Name", "Unknown"))

            # Build steps (skip PDF if ReportLab missing)
            steps = ["insert_db", "followup_rule", "tag_status"]
            if REPORTLAB_OK:
                steps.append("generate_pdf")
            steps.append("write_json")

            # Launch new real-time dialog
            ctx = {"data": dict(self.current_data)}
            dlg = AgentRunDialog(self.agent, steps, ctx, self)
            dlg.exec_()

        except Exception as e:
            QtWidgets.QMessageBox.critical(self, self.tr("Agent Error"), self.tr("The Agent run failed:\n") + str(e))


# ---------- Standalone ----------
if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    try:
        from UI.design_system import apply_global_theme, apply_window_backdrop
        apply_global_theme(app, base_point_size=11)
    except Exception:
        app.setStyle("Fusion")
    w = ExtractionTab(); w.resize(1100, 720); w.show()
    try:
        from UI.design_system import apply_window_backdrop
        apply_window_backdrop(w, prefer_mica=True)
    except Exception:
        pass
    sys.exit(app.exec_())
